import pytest
import time

from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By


from pageObjects.LoginPage import LoginPage
from pageObjects.AddcustomerPage import AddCustomer
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen
import string
import random

class Test_003_AddCustomer:
    baseURL = ReadConfig.getApplicationURL()
    username = ReadConfig.getUseremail()
    password = ReadConfig.getPassword()
    logger = LogGen.loggen()  # Logger
    path = ".//TestData/LoginData.xlsx"

    @pytest.mark.sanity
    @pytest.mark.regression
    def test_addCustomer(self,setup):
        self.logger.info("************* Test_003_AddCustomer **********")
        self.driver=setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting Add Customer Test **********")

        self.addcust = AddCustomer(self.driver)
        self.addcust.clickOnCustomersMenu()
        self.addcust.clickOnCustomersMenuItem()

        self.addcust.clickOnAddnew()

        self.logger.info("************* Providing customer info **********")

        self.email = random_generator() + "@gmail.com"
        self.addcust.setEmail(self.email)
        email = self.email
        print(email)
        time.sleep(2)

        workbook = load_workbook(self.path)
        sheet = workbook.active
        # Find the last row in the sheet
        last_row = sheet.max_row

        # Check if the cell contains data
        cell_value = sheet['A1'].value

        if cell_value:
            # If cell contains data, clear the cell and store the new data
            sheet['B1'].value = None
            sheet['C1'].value = None
            sheet['B1'] = email
        else:
            # If cell is empty, store the generated email
            sheet['B1'] = email

        # Save the workbook
        workbook.save("TestData/LoginData.xlsx")



        # self.addcust.setPassword("test123")
        # self.addcust.setCustomerRoles("Guests")
        # self.addcust.setManagerOfVendor("Vendor 2")
        # self.addcust.setGender("Male")
        # self.addcust.setFirstName("Pavan")
        # self.addcust.setLastName("Kumar")
        # self.addcust.setDob("7/05/1985")  # Format: D / MM / YYY
        # self.addcust.setCompanyName("busyQA")
        # self.addcust.setAdminContent("This is for testing.........")
        # self.addcust.clickOnSave()
        #
        # self.logger.info("************* Saving customer info **********")
        #
        # self.logger.info("********* Add customer validation started *****************")
        #
        # # Locate the body element and retrieve its text content
        # body_element = self.driver.find_element(By.TAG_NAME, "body")
        # self.msg = body_element.text
        #
        # # Print the text content of the body element
        # print(self.msg)
        #
        # if 'customer has been added successfully.' in self.msg:
        #     assert True
        #     self.logger.info("********* Add customer Test Passed *********")
        # else:
        #     self.driver.save_screenshot(".\\Screenshots\\" + "test_addCustomer_scr.png")  # Screenshot
        #     self.logger.error("********* Add customer Test Failed ************")
        #     assert False
        #
        # self.driver.close()
        # self.logger.info("******* Ending Add customer test **********")


def random_generator(size=8, chars=string.ascii_lowercase + string.digits):
        return ''.join(random.choice(chars) for x in range(size))

