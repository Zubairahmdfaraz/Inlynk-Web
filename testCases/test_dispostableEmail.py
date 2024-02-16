import re
import time
import unittest

from selenium.webdriver.support import wait
from selenium.webdriver.support import expected_conditions as EC
import time
import re

from selenium.webdriver.support.wait import WebDriverWait

from pageObjects.ConfigurationPage import ConfigurationPage
from pageObjects.AddEmployeesPage import AddEmployeesPage
import pytest
from openpyxl.reader.excel import load_workbook
from selenium.common import NoSuchElementException, StaleElementReferenceException, TimeoutException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from pageObjects.LoginPage import LoginPage
from testCases.conftest import setup
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig
from pageObjects.companySignUpPage import companySignUpPage
from pageObjects.randomGen import randomGen
from selenium import webdriver


class TestSignUpEmailTry(unittest.TestCase):
    baseURL = ReadConfig.getApplicationURL()
    setSearchIndustryType = "Information Technology"
    password = ReadConfig.getPassword()

    DeptName = "Emp creation QA"
    DeptDescription = "Emp creation Software Testing"

    def setUp(self):
        self.logger = LogGen.loggen()
        self.driver = webdriver.Chrome()  # Change to the appropriate driver
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)

    def tearDown(self):
        self.driver.quit()

    @pytest.mark.run(order=1)
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    @pytest.mark.regression
    def test_SignUpwithValid(self):
        self.driver.get(self.baseURL)
        self.logger.info("******** Starting test_Sign Up with Valid ***********")
        self.logger.info("******** User is on Login page ***********")

        email = randomGen.random_email()
        first_name = randomGen.random_first_name()
        company_name = randomGen.random_company_name()
        phone_number = randomGen.random_phone_number()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['I10'] = email
        ws['H10'] = company_name

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.clickCompanysignupButton()
        self.logger.info("******** user is in company signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        self.sp.setCompanyName(company_name)

        self.sp.setSearchIndustryType(self.setSearchIndustryType)
        self.sp.selectCompany()
        self.sp.setContactName(first_name)
        self.sp.setEmail(email)
        self.sp.clickcountrydd()
        self.sp.clickindia()
        self.sp.clickstatedd()
        self.sp.clickTelangana()
        self.sp.clickcitydd()
        self.sp.clickHyderabad()
        self.sp.setPhone(phone_number)
        self.sp.setPassword(self.password)
        self.sp.setConfirmPassword(self.password)
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** user navigated to enter OTP page ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("http://mailcatch.com/en/disposable-email")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@name='box']")
        yopmail.send_keys(email + Keys.ENTER)
        time.sleep(1)

        reload_button = self.driver.find_element(By.XPATH, "//img[@title='Reload']")

        # Click the Reload button every second until the subject is displayed or a maximum time is reached
        max_wait_time = 60  # Set your maximum wait time in seconds
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            reload_button.click()

            try:
                # Check if the subject is displayed
                subject = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                )
                subject.click()
                break  # Break out of the loop if subject is displayed
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred. Retrying...")
                continue  # Retry the loop if StaleElementReferenceException occurs
            except TimeoutException:
                time.sleep(0.1)

        iframeElement = self.driver.find_element(By.ID, "emailframe")
        self.driver.switch_to.frame(iframeElement)

        # Code outside the loop will be executed after the loop or when a TimeoutException occurs
        otp = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        # This code is for QA ENV
        otp = self.driver.find_element(By.XPATH,"//body")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])

        self.sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info("******** Verifying the OTP ***********")
        self.sp.clickVerifyButton()
        self.sp.clickContinueToLogin()
        self.logger.info("******** Company Sign Up successful ***********")
        self.logger.info("******** Entering the sig up credentials for Login ***********")
        # Read data from specific cells
        email = ws['I10'].value

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()

        self.logger.info("****Started Create New Employee in Super Admin and Admin Account ****")
        first_name = randomGen.random_first_name()
        email = randomGen.random_email()
        phone_number = randomGen.random_phone_number()
        Emp_Id = randomGen.random_Emp_Id()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['J10'] = first_name
        ws['K10'] = "personal" + email


        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.aep.clickActive()
        time.sleep(2)
        self.aep.clickNewButton()
        self.aep.setFullname(first_name)

        self.aep.setEmail("emp" + email)
        self.aep.setPersonalEmail("personal" + email)

        self.aep.setPhoneNumber(phone_number)

        self.aep.setEmpId(Emp_Id)
        self.logger.info("**** Started Create  New Department in Super Admin Account****")
        self.aep.clickAddDeptButton()
        self.cp = ConfigurationPage(self.driver)
        self.cp.setDepartmentName("Department " + first_name)
        self.cp.setEnterDescription(self.DeptDescription)
        self.aep.clickDoneAddDept()
        time.sleep(3)
        act_Text = self.cp.Text_DeptCreatedSuccessful()
        if act_Text == "Department created successfully":
            assert True
            self.logger.info("********* Create Department Test is Passed ***********")

        else:
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_createDept.png")
            self.logger.error("********* Create Department Test is Failed ***********")
            self.driver.close()
            assert False

        self.logger.info("**** Started Create  New Division in Super Admin Account****")
        self.aep.clickAddDivisionButton()
        self.aep.setDivisionName("Division " + first_name)
        self.cp.setEnterDescription(self.DeptDescription)
        self.aep.clickDoneAddDept()
        time.sleep(3)
        act_Text = self.cp.Text_DivisionCreatedSuccessful()
        if act_Text == "Division created successfully":
            assert True
            self.logger.info("********* Create Division Test is Passed ***********")

        else:
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_createDept.png")
            self.logger.error("********* Create Division Test is Failed ***********")
            self.driver.close()
            assert False

        self.logger.info("**** Started Create  New Designation in Super Admin Account****")
        self.aep.clickAddDesignation()
        self.aep.setDivisionName("Designation " + first_name)
        self.cp.setEnterDescription(self.DeptDescription)
        self.aep.clickDoneAddDept()
        time.sleep(3)
        act_Text = self.cp.Text_DesignationCreatedSuccessful()
        if act_Text == "Designation created successfully":
            assert True
            self.logger.info("********* Create Designation Test is Passed ***********")
        else:
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_createDept.png")
            self.logger.error("********* Create Designation Test is Failed ***********")
            self.driver.close()
            assert False

        # time.sleep(2)
        self.aep.clickCountryDD()
        # time.sleep(3)
        self.sp = companySignUpPage(self.driver)
        self.sp.clickindia()
        self.sp.clickstatedd()
        self.sp.clickTelangana()
        self.sp.clickcitydd()
        self.sp.clickHyderabad()
        time.sleep(2)
        self.aep.clickAddButton()
        time.sleep(6)
        act_Text = self.aep.Text_EmployeeCreatedSuccessful()
        if act_Text == "Employee created successfully":
            assert True
            self.logger.info("********* Employee creation Test is Passed ***********")

        else:
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_createDept.png")
            self.logger.error("********* Employee creation Test is Failed ***********")
            self.driver.close()
            assert False

    if __name__ == '__main__':
        unittest.main(verbosity=2)


if __name__ == '__main__':
    unittest.main(verbosity=2)
