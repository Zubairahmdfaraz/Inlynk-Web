import time
from telnetlib import EC

import pytest
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from selenium.common import NoSuchElementException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

from pageObjects.LoginPage import LoginPage
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig
from pageObjects.companySignUpPage import companySignUpPage
from pageObjects.randomGen import randomGen
from selenium import webdriver
from openpyxl import workbook


class TestInividualSignUp:
    baseURL = ReadConfig.getApplicationURL()
    password = ReadConfig.getPassword()

    # @pytest.fixture(scope="class")
    # def setup(self):
    #     self.driver = webdriver.Chrome()  # Change to the appropriate driver
    #     self.driver.maximize_window()
    #     yield
    #     self.driver.quit()

    #     @pytest.mark.run(order=1)
    #     def test_IndividualSignUpwithValid(self, setup):
    #         self.logger = LogGen.loggen()
    #         self.logger.info("****Opening URL****")
    #         self.driver = setup
    #         self.driver.maximize_window()
    #         self.driver.get(self.baseURL)
    #         self.logger.info("******** Starting test_Individual Sign Up with Valid ***********")
    #         self.logger.info("******** User is on Login page ***********")
    #
    #         email = randomGen.random_email()
    #         first_name = randomGen.random_first_name()
    #         phone_number = randomGen.random_phone_number()
    #
    #         self.logger.info("******** Generating and storing data into excel sheet ***********")
    #         # Load the existing workbook
    #         wb = load_workbook("TestData/LoginData.xlsx")
    #
    #         # Select the active worksheet
    #         ws = wb.active
    #
    #         # Update the existing cells with new data
    #         ws['A4'] = email
    #         ws['B4'] = first_name
    #         ws['C4'] = self.password
    #         ws['D4'] = phone_number
    #
    #         # Save the workbook
    #         wb.save("TestData/LoginData.xlsx")
    #
    #         self.sp = companySignUpPage(self.driver)
    #         self.sp.clicksignuplink()
    #         self.sp.clickIndividualSignupButton()
    #         self.logger.info("******** user is in Individual signup page ***********")
    #         self.logger.info("******** Entering valid data into the fields ***********")
    #
    #         self.sp.setFullName("IE"+first_name)
    #         self.sp.setEmail(email)
    #         self.sp.setPhone(phone_number)
    #         self.sp.setPassword(self.password)
    #         self.sp.setConfirmPassword(self.password)
    #         self.sp.clicktermsConditions()
    #         time.sleep(2)
    #         self.logger.info("******** Clicking on signup button ***********")
    #         self.logger.info("******** user navigated to enter OTP page ***********")
    #         self.sp.clicksignupNow()
    #         time.sleep(2)
    #
    # # Execute JavaScript to open a new tab
    #         self.driver.execute_script("window.open('about:blank', '_blank');")
    #
    #         # Perform actions in the new tab (if needed)
    #         # For example:
    #         self.driver.switch_to.window(self.driver.window_handles[1])
    #         self.logger.info("******** Opening new url in another tab for Email OTP ***********")
    #         time.sleep(1)
    #         self.driver.get("https://yopmail.com/")
    #         time.sleep(1)
    #         yopmail = self.driver.find_element(By.XPATH, "//input[@id='login']")
    #         yopmail.send_keys(email + Keys.ENTER)
    #         time.sleep(1)
    #         iframeElement = self.driver.find_element(By.ID, "ifmail")
    #         self.driver.switch_to.frame(iframeElement)
    #
    #         # here appears the captcha
    #         captcha_retries = 3
    #         attempts = 0
    #         captcha_found = False
    #         while attempts < captcha_retries:
    #             try:
    #                 # Your existing code for handling the CAPTCHA
    #                 # ...
    #
    #
    #                 # If CAPTCHA is successfully filled, set captcha_found to True
    #                 captcha_found = False  # Update this line accordingly
    #
    #                 # Verify the presence of specific text
    #                 text_element = self.driver.find_element(By.XPATH, "//div[@class='b alc r_message']")
    #                 if text_element.text == "Complete the CAPTCHA to continue":
    #                     captcha_found = True
    #                     break  # Exit the loop if expected text is found
    #
    #             except NoSuchElementException:
    #                 # No CAPTCHA found, continue with the rest of the steps
    #                 captcha_found = False  # Update this line accordingly
    #                 break
    #
    #             except Exception as e:
    #                 print(f"Error occurred: {e}")
    #                 attempts += 1
    #                 time.sleep(2)
    #
    #             if captcha_found:
    #                 print("Expected text found. Rerunning the test.")
    #                 self.test_IndividualSignUpwithValid(setup)  # Rerun the test method
    #
    #             else:
    #                 print("Expected text not found. Continuing with the code.")
    #                 # Continue with the remaining steps of your test script
    #                 time.sleep(2)  # Add any necessary wait times or other steps
    #
    #         # iframeElement = self.driver.find_element(By.ID, "ifmail")
    #         # self.driver.switch_to.frame(iframeElement)
    #         self.logger.info("******** Email Copied ***********")
    #         time.sleep(1)
    #         otp = self.driver.find_element(By.XPATH,
    #                                        "/html[1]/body[1]/main[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/h1[1]")
    #         self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
    #         time.sleep(0.5)
    #         getOTP = otp.text
    #         print(getOTP)
    #         self.logger.info("******** Switching back and entering the otp ***********")
    #         self.driver.switch_to.default_content()
    #
    #         self.driver.switch_to.window(self.driver.window_handles[0])
    #
    #         self.sp.setOtp(getOTP)
    #
    #         time.sleep(2)
    #         self.logger.info("******** Verifying the OTP ***********")
    #         self.sp.clickVerifyButton()
    #         self.sp.clickContinueToLogin()
    #         self.logger.info("******** Individual Sign Up successful ***********")
    #         self.logger.info("******** Entering the sig up credentials for Login ***********")
    #
    #         # Read data from specific cells
    #         email = ws['A4'].value
    #
    #         self.lp = LoginPage(self.driver)
    #         self.lp.setUserName(email)
    #         self.lp.setPassword(self.password)
    #         self.lp.clickLogin()
    #         self.lp.clickcreatePost()
    #         self.logger.info("******** Login successful ***********")
    #         act_Text = self.lp.newsFeedText()
    #
    #         if act_Text == "Create News Feed":
    #             assert True
    #             self.logger.info("********* Individual SignUp Test ith valid data is Passed ***********")
    #
    #         else:
    #             self.driver.save_screenshot(".\\ScreenShots\\" + "test_IndividualSignUpwithValid.png")
    #             self.logger.error("********* Individual SignUp Test is Failed ***********")
    #             self.driver.close()
    #             assert False
    #
    #         time.sleep(3)
    #         self.driver.find_element(By.XPATH, "//div[@class='flexAutoRow alignCntr pdngHXS']").click()
    #         self.driver.close()

    @pytest.mark.flaky(reruns=3)  # Decorate the test with flaky, specifying the maximum number of runs
    @pytest.mark.run(order=2)
    def test_IndividualSignUpWith_inValid_Data(self, setup):
        self.logger = LogGen.loggen()
        self.logger.info("****Opening URL****")
        self.driver = setup
        self.driver.maximize_window()
        self.driver.get(self.baseURL)
        self.logger.info("******** Starting test_Individual Sign Up with Valid ***********")
        self.logger.info("******** User is on Login page ***********")

        email = "123sdd"
        first_name = "11111111444"
        phone_number = "8888888888"
        Password = "Inlink1223"

        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.clickIndividualSignupButton()
        self.logger.info("******** user is in Individual signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")

        self.sp.setFullName(first_name)
        self.sp.setEmail(email)
        self.sp.setPhone(phone_number)
        self.sp.setPassword(Password)
        self.sp.setConfirmPassword(self.password)
        time.sleep(2)
        self.sp.clicktermsConditions()
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # List of error texts to validate
        error_texts = [
            "Please enter valid company name",
            "Enter valid email address",
            "The password must be 8 characters in length, and must contains at least one small letter,one capital letter, one numeric character and one special character",
            "Password and confirm password are not same"
            # Add more error texts if needed
        ]

        # Find elements containing error texts and validate
        for text in error_texts:
            error_elements = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")

            if error_elements:
                self.logger.info(f"Found error text: {text}")
                assert True
                # self.logger.info("********* Validation of Error message with invalid data test is Passed ***********")
                # Additional actions can be performed here if needed
            else:
                self.logger.info(f"Error text not found: {text}")
                self.driver.save_screenshot(".\\ScreenShots\\" + "test_IndividualSignUpWith_inValid_Data.png")
                # self.logger.error("********* Validation of Error message with invalid data test is Failed ***********")
                self.driver.close()
                assert False
