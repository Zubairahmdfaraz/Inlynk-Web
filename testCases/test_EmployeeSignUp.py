import time
# from telnetlib3 import EC


import pytest
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from selenium.common import NoSuchElementException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from sunithaPageObjects.MyProfile import MyprofilePage
from pageObjects.LoginPage import LoginPage
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig
from pageObjects.companySignUpPage import companySignUpPage
from pageObjects.randomGen import randomGen
from pageObjects.AddEmployeesPage import AddEmployeesPage
from pageObjects.EmployeeModulePage import EmployeeModulePage
from selenium import webdriver
from openpyxl import workbook
from selenium.webdriver.support import expected_conditions as EC
class TestEmployeeSignUp:
    baseURL = ReadConfig.getApplicationURL()
    password = ReadConfig.getPassword()




    @pytest.mark.run(order=1)
    # @pytest.mark.parametrize("run_number", range(1, 2))
    @pytest.mark.skip(reason="skip for now")
    # def test_EmployeeSignUpWithValid(self, run_number, setup):
    def test_EmployeeSignUpValidWithoutDomain(self, setup):
        self.logger = LogGen.loggen()
        self.logger.info("****Opening URL****")
        self.driver = setup
        self.driver.maximize_window()
        self.driver.get(self.baseURL)
        self.logger.info("******** Starting test_Employee Sign Up with Valid ***********")
        self.logger.info("******** User is on Login page ***********")

        email = randomGen.random_email()
        first_name = randomGen.random_first_name()
        phone_number = randomGen.random_phone_number()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        companyName = ws['A6'].value
        companyEmail = ws['B6'].value

        # Update the existing cells with new data
        ws['E6'] = email
        ws['C6'] = first_name
        ws['D6'] = phone_number

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")
        # Verify the Signup functionality. with positive data.
        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.ClickEmployeeSignUp()
        self.logger.info("******** user is in company signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        # 1. Verify the functionality and accuracy of the Company Selection dropdown.
        self.sp.setSearchCompany(companyName)
        self.sp.ClickSelectCompany()
        self.sp.setFullName(first_name)
        self.sp.setEmail(email)
        self.sp.setPhone(phone_number)
        self.sp.setPassword(self.password)
        self.sp.setConfirmPassword(self.password)
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** user navigated to enter OTP page ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("https://yopmail.com/")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@id='login']")
        yopmail.send_keys(email + Keys.ENTER)
        time.sleep(1)
        iframeElement = self.driver.find_element(By.ID, "ifmail")
        self.driver.switch_to.frame(iframeElement)

        # here appears the captcha
        captcha_retries = 3
        attempts = 0
        captcha_found = False
        while attempts < captcha_retries:
            try:
                # Your existing code for handling the CAPTCHA
                # ...

                # If CAPTCHA is successfully filled, set captcha_found to True
                captcha_found = False  # Update this line accordingly

                # Verify the presence of specific text
                text_element = self.driver.find_element(By.XPATH, "//div[@class='b alc r_message']")
                if text_element.text == "Complete the CAPTCHA to continue":
                    captcha_found = True
                    break  # Exit the loop if expected text is found

            except NoSuchElementException:
                # No CAPTCHA found, continue with the rest of the steps
                captcha_found = False  # Update this line accordingly
                break

            except Exception as e:
                print(f"Error occurred: {e}")
                attempts += 1
                time.sleep(2)

            if captcha_found:
                print("Expected text found. Rerunning the test.")
                self.test_EmployeeSignUpValidWithoutDomain(setup)  # Rerun the test method

            else:
                print("Expected text not found. Continuing with the code.")
                # Continue with the remaining steps of your test script
                time.sleep(2)  # Add any necessary wait times or other steps

        # iframeElement = self.driver.find_element(By.ID, "ifmail")
        # self.driver.switch_to.frame(iframeElement)
        self.logger.info("******** Email Copied ***********")
        time.sleep(1)
        otp = self.driver.find_element(By.XPATH,
                                       "/html[1]/body[1]/main[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/h1[1]")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)
        getOTP = otp.text
        print(getOTP)
        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])

        self.sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info("******** Verifying the OTP ***********")
        self.sp.clickVerifyButton()
        self.sp.clickContinueToLogin()
        # self.logger.info("******** Employee Sign Up successful ***********")
        # self.logger.info("******** Entering the sig up credentials for Login ***********")
        # # Read data from specific cells
        # email = ws['A2'].value
        #
        # self.lp = LoginPage(self.driver)
        # self.lp.setUserName(email)
        # self.lp.setPassword(self.password)
        # self.lp.clickLogin()
        # self.lp.clickcreatePost()
        # self.logger.info("******** Login successful ***********")
        # act_Text = self.lp.newsFeedText()
        #
        # if act_Text == "Create News Feed":
        #     assert True
        #     self.logger.info("********* Employee SignUp Test is Passed ***********")
        #
        # else:
        #     self.driver.save_screenshot(".\\ScreenShots\\" + "test_EmployeeSignUpwithValid.png")
        #     self.logger.error("********* Employee SignUp Test is Failed ***********")
        #     self.driver.close()
        #     assert False
        #
        # time.sleep(3)
        # self.driver.find_element(By.XPATH, "//div[@class='flexAutoRow alignCntr pdngHXS']").click()
        # self.driver.close()

    @pytest.mark.run(order=2)
    @pytest.mark.skip(reason="Skipping this test")
    def test_EmployeeSignUpWith_inValid_Data(self, setup):
        self.logger = LogGen.loggen()
        self.logger.info("****Opening URL****")
        self.driver = setup
        self.driver.maximize_window()
        self.driver.get(self.baseURL)
        self.logger.info("******** Starting test_Employee Sign Up with Valid ***********")
        self.logger.info("******** User is on Login page ***********")

        email = "123sdd"
        first_name = "11111111444"
        phone_number = "8888888888"
        Password = "Inlink1223"

        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.ClickEmployeeSignUp()
        self.logger.info("******** user is in Employee signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        companyName = ws['A6'].value
        self.sp.setSearchCompany(companyName)
        self.sp.ClickSelectCompany()
        self.sp.setFullName(first_name)
        self.sp.setEmail(email)
        self.sp.setPhone(phone_number)
        self.sp.setPassword(Password)
        self.sp.setConfirmPassword(self.password)
        time.sleep(2)
        self.sp.clicktermsConditions()
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # List of error texts to validate
        error_texts = [
            "Please enter valid fullName",
            "Email must start with alphabet",
            "The password must be 8 characters in length, and must contains at least one small letter,one capital letter, one numeric character and one special character",
            "Password and confirm password are not same"
            # Add more error texts if needed
        ]

        # Find elements containing error texts and validate
        for text in error_texts:
            error_elements = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")

            if error_elements:
                self.logger.info(f"Found error text: {text}")
                assert True
                # self.logger.info("********* Validation of Error message with invalid data test is Passed ***********")
                # Additional actions can be performed here if needed
            else:
                self.logger.info(f"Error text not found: {text}")
                self.driver.save_screenshot(".\\ScreenShots\\" + "test_EmployeeSignUpWith_inValid_Data.png")
                # self.logger.error("********* Validation of Error message with invalid data test is Failed ***********")
                self.driver.close()
                assert False

    @pytest.mark.run(order=4)
    @pytest.mark.skip(reason="skip for now")
    def test_ApproveSignedUpEmployee(self, setup):
        # to use same class different test method
        self.test_EmployeeSignUpValidWithoutDomain(setup)
        # Create an instance of LoginTestClass
        # login_test_instance = LoginTestClass()
        # # Call the test method from LoginTestClass
        # login_test_instance.test_EmployeeSignUpWith_inValid_Data(setup)

        # self.logger = LogGen.loggen()
        # self.logger.info("****Opening URL****")
        # self.driver = setup
        # self.driver.maximize_window()
        # self.driver.get(self.baseURL)
        # self.logger.info("******** Starting test_Employee Sign Up with Valid ***********")
        # self.logger.info("******** User is on Login page ***********")
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active
        email = ws['B6'].value
        first_name = ws['C6'].value
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email)
        self.lp.setPassword("Inlink@123")
        # self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.em = EmployeeModulePage(self.driver)
        self.em.ClickPendingTab()
        self.em.setPendingSearchField(first_name)
        self.em.ClickStatusDD()
        self.em.ClickStatusApprove()
        Emp_Id = randomGen.random_Emp_Id()
        self.em.setEmpId(Emp_Id)
        self.em.ClickDepartmentDD()
        self.em.ClickSelectDD()
        self.em.ClickDivisionDD()
        self.em.ClickSelectDD()
        self.em.ClickDesignationDD()
        self.em.ClickSelectDD()
        self.em.ClickApproveButton()
        time.sleep(3)
        Verify_texts = [
            "Employee approved successfully"
            # Add more error texts if needed
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            error_elements = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")

            if error_elements:
                self.logger.info(f"Found Toast Message: {text}")
                assert True
                # self.logger.info("********* Validation of Error message with invalid data test is Passed ***********")
                # Additional actions can be performed here if needed
            else:
                self.logger.info(f"Toast Message not found: {text}")
                self.driver.save_screenshot(".\\ScreenShots\\" + "test_ApproveSignedUpEmployee.png")
                # self.logger.error("********* Validation of Error message with invalid data test is Failed ***********")
                self.driver.close()
                assert False

        # Verify active tab
        self.logger.info("Verifying the Approved Employee in Active Tab")
        self.aep.clickActive()
        time.sleep(1)
        self.aep.setActiveSearchField(first_name)
        element = self.driver.find_element(By.XPATH, "//span[contains(text(),'" + first_name + "')]")
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_ApproveSignedUpEmployee.png")
            self.driver.close()
            assert False

    @pytest.mark.run(order=3)
    @pytest.mark.skip(reason="skip for now")
    def test_RejectSignedUpEmployee(self, setup):
        # to use same class different test method
        self.test_EmployeeSignUpValidWithoutDomain(setup)
        # Create an instance of LoginTestClass
        # login_test_instance = LoginTestClass()
        # # Call the test method from LoginTestClass
        # login_test_instance.test_EmployeeSignUpWith_inValid_Data(setup)

        # self.logger = LogGen.loggen()
        # self.logger.info("****Opening URL****")
        # self.driver = setup
        # self.driver.maximize_window()
        # self.driver.get(self.baseURL)
        # self.logger.info("******** Starting test_Employee Sign Up with Valid ***********")
        # self.logger.info("******** User is on Login page ***********")
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active
        email = ws['B6'].value
        first_name = ws['C6'].value
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email)
        self.lp.setPassword("Inlink@123")
        # self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.em = EmployeeModulePage(self.driver)
        self.em.ClickPendingTab()
        self.em.setPendingSearchField(first_name)
        self.em.ClickStatusDD()
        self.em.ClickStatusReject()
        self.em.ClickConfReject()
        time.sleep(5)
        Verify_texts = [
            "Employee rejected successfully"
            # Add more error texts if needed
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            error_elements = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")

            if error_elements:
                self.logger.info(f"Found Toast Message: {text}")
                assert True
                # self.logger.info("********* Validation of Error message with invalid data test is Passed ***********")
                # Additional actions can be performed here if needed
            else:
                self.logger.info(f"Toast Message not found: {text}")
                self.driver.save_screenshot(".\\ScreenShots\\" + "test_RejectSignedUpEmployee.png")
                # self.logger.error("********* Validation of Error message with invalid data test is Failed ***********")
                self.driver.close()
                assert False

        # Verify active tab
        self.logger.info("Verifying the rejected Employee in rejected Tab")
        self.em.ClickRejectTab()
        time.sleep(1)
        self.em.setRejectedSearchField(first_name)
        element = self.driver.find_element(By.XPATH, "//span[contains(text(),'" + first_name + "')]")
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_ApproveSignedUpEmployee_Valid_Data.png")
            self.driver.close()
            assert False

    @pytest.mark.smoke
    @pytest.mark.skip(reason="skip for now")
    @pytest.mark.run(order=5)
    def test_EmployeeSignUpValidWithDomain(self, setup):
        self.logger = LogGen.loggen()
        self.logger.info("****Opening URL****")
        self.driver = setup
        self.driver.maximize_window()
        self.driver.get(self.baseURL)
        self.logger.info("******** Starting test_Employee Sign Up with Valid ***********")
        self.logger.info("******** User is on Login page ***********")

        # email = randomGen.random_email()
        first_name = randomGen.random_first_name()
        phone_number = randomGen.random_phone_number()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        companyName = ws['A10'].value
        companyEmail = ws['B10'].value

        # Update the existing cells with new data
        # ws['E10'] = email
        ws['C10'] = first_name
        ws['D10'] = phone_number

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")
        # Verify the Signup functionality. with positive data.
        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.ClickEmployeeSignUp()
        self.logger.info("******** user is in company signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        # 1. Verify the functionality and accuracy of the Company Selection dropdown.
        self.sp.setSearchCompany(companyName)
        self.sp.ClickSelectCompany()
        self.sp.setFullName(first_name)
        time.sleep(0.5)

        self.sp.setEmail(first_name + "20")
        Email = first_name + "20@yopmail.com"
        ws['E10'] = Email
        wb.save("TestData/LoginData.xlsx")
        time.sleep(1)
        self.sp.setPhone(phone_number)
        self.sp.setPassword(self.password)
        self.sp.setConfirmPassword(self.password)
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** user navigated to enter OTP page ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("https://yopmail.com/")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@id='login']")
        yopmail.send_keys(first_name + "20@yopmail.com" + Keys.ENTER)
        time.sleep(1)
        iframeElement = self.driver.find_element(By.ID, "ifmail")
        self.driver.switch_to.frame(iframeElement)

        # here appears the captcha
        captcha_retries = 3
        attempts = 0
        captcha_found = False
        while attempts < captcha_retries:
            try:
                # Your existing code for handling the CAPTCHA
                # ...

                # If CAPTCHA is successfully filled, set captcha_found to True
                captcha_found = False  # Update this line accordingly

                # Verify the presence of specific text
                text_element = self.driver.find_element(By.XPATH, "//div[@class='b alc r_message']")
                if text_element.text == "Complete the CAPTCHA to continue":
                    captcha_found = True
                    break  # Exit the loop if expected text is found

            except NoSuchElementException:
                # No CAPTCHA found, continue with the rest of the steps
                captcha_found = False  # Update this line accordingly
                break

            except Exception as e:
                print(f"Error occurred: {e}")
                attempts += 1
                time.sleep(2)

            if captcha_found:
                print("Expected text found. Rerunning the test.")
                self.test_EmployeeSignUpValidWithDomain(setup)  # Rerun the test method

            else:
                print("Expected text not found. Continuing with the code.")
                # Continue with the remaining steps of your test script
                time.sleep(2)  # Add any necessary wait times or other steps

        # iframeElement = self.driver.find_element(By.ID, "ifmail")
        # self.driver.switch_to.frame(iframeElement)
        self.logger.info("******** Email Copied ***********")
        time.sleep(1)
        otp = self.driver.find_element(By.XPATH,
                                       "/html[1]/body[1]/main[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/h1[1]")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)
        getOTP = otp.text
        print(getOTP)
        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])

        self.sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info("******** Verifying the OTP ***********")
        self.sp.clickVerifyButton()
        self.sp.clickContinueToLogin()

        # self.driver.close()

    @pytest.mark.run(order=6)
    @pytest.mark.smoke
    # @pytest.mark.flaky(rerun=3, rerun_delay=2)
    @pytest.mark.skip(reason="skip for now")
    def test_ActiveSignedUpEmployeeWithDomain(self, setup):
        # to use same class different test method
        self.test_EmployeeSignUpValidWithDomain(setup)

        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active
        email = ws['B10'].value
        first_name = ws['C10'].value
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email)
        self.lp.setPassword("Inlink@123")
        # self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.em = EmployeeModulePage(self.driver)

        self.aep.clickActive()
        time.sleep(2)
        self.aep.setActiveSearchField(first_name)
        element = self.driver.find_element(By.XPATH, "//span[contains(text(),'" + first_name + "')]")
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_ApproveSignedUpEmployee.png")
            self.driver.close()
            assert False

        element.click()

    @pytest.mark.run(order=7)
    @pytest.mark.tests
    # @pytest.mark.flaky(rerun=3, rerun_delay=2)
    # @pytest.mark.skip(reason="skip for now")
    def test_AdminSignedUpEmployeeWithDomain(self, setup):
        self.test_ActiveSignedUpEmployeeWithDomain(setup)
        self.aep.ClickEmployeeStatus()
        self.aep.ClickAdminStatus()
        self.aep.ClickGrantAdmin()
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active
        first_name = ws['C10'].value

        xpath = "//div[contains(text(), '" + first_name + " is an admin now')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "AdminSignedUpEmployeeWithDomain.png")
            self.driver.close()
            self.driver.quit()
            assert False

        self.lp.clickLogout()
        self.logger.info("******** Entering Admin Employee credentials for verification ***********")
        # Read data from specific cells
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        Username = ws['E10'].value
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(Username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.my = MyprofilePage(self.driver)
        self.my.clickMyProfileModule()

        adminxpath = "// span[text() = 'admin']"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, adminxpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee Role : {element.text}")
            assert True
            self.driver.quit()
        else:
            self.logger.info(f"Employee Role not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "VerifyAdminRole.png")
            self.driver.close()
            self.driver.quit()
            assert False