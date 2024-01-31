import unittest

import pytest
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from pageObjects.networksPage import networksPage

from pageObjects.LoginPage import LoginPage
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen


class TestLogin(unittest.TestCase):
    baseURL = ReadConfig.getApplicationURL()
    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    username = worksheet["A2"].value
    RMname = worksheet["A8"].value
    password = ReadConfig.getPassword()
    CompanyManufacture = worksheet["H2"].value
    EmailManufacture = worksheet["I2"].value
    CompanyPartner = worksheet["H3"].value
    EmailPartner = worksheet["I3"].value
    CompanyDistributor = worksheet["H4"].value
    EmailDistributor = worksheet["I4"].value
    CompanyVendor = worksheet["H5"].value
    EmailVendor = worksheet["I5"].value

    workbook.close()

    logger = LogGen.loggen()

    def setUp(self):
        self.logger = LogGen.loggen()
        self.driver = webdriver.Chrome()  # Change to the appropriate driver
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)
        self.logger.info("****Opening URL****")
        self.driver.get(self.baseURL)

    def tearDown(self):
        self.driver.quit()

    @pytest.mark.smoke(order=1)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_ConnectionCompanyAsManufacturer(self):
        self.logger.info("****Started Network Connection Test****")
        self.lp = LoginPage(self.driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + self.username + " and password " + self.password)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)

        self.lp.clickLogin()
        self.np = networksPage(self.driver)
        self.np.clickNetworks()
        self.np.setsearchField(self.CompanyManufacture)
        # Verifying and Clicking on the company
        self.driver.find_element(By.XPATH, "//span[contains(text(),'"+self.CompanyManufacture+"')]").click()
        self.np.clickConnectButton()
        self.np.clickDropDownList()
        self.np.clickManufacturer()
        self.np.setRM_searchField(self.RMname)
        self.np.clickSelectRM()
        self.np.clickcheckbox()
        self.np.clickButtonConnect2()
        # List of error texts to validate
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            error_elements = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")

            if error_elements:
                self.logger.info(f"Found error text: {text}")
                assert True

            else:
                self.logger.info(f"Error text not found: {text}")
                self.driver.save_screenshot(".\\ScreenShots\\" + "test_ConnectionCompanyAsManufacturer.png")
                self.driver.close()
                assert False

        self.np.clickOKButton()
        self.lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        # Read data from specific cells
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmail = ws['I2'].value
        companyName = ws['C2'].value

        self.lp = LoginPage(self.driver)
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + CompEmail + " and Password:" + self.password)
        self.lp.setUserName(CompEmail)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.np.clickNetworks()
        self.np.clickPendingTab()
        self.np.setsearchField(companyName)
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + companyName + "')]").click()
        self.np.clickApproveButton()
        self.np.clickAcceptButton()

        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            error_elements = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")

            if error_elements:
                self.logger.info(f"Found error text: {text}")
                assert True

            else:
                self.logger.info(f"Error text not found: {text}")
                self.driver.save_screenshot(".\\ScreenShots\\" + "test_ConnectionCompanyAsManufacturer.png")
                self.driver.close()
                assert False





    if __name__ == '__main__':
        unittest.main(verbosity=2)
