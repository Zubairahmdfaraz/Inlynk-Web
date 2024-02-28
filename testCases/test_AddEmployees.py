import time
import unittest

import pytest
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pageObjects.ConfigurationPage import ConfigurationPage
from pageObjects.LoginPage import LoginPage
from pageObjects.AddEmployeesPage import AddEmployeesPage
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen
from pageObjects.randomGen import randomGen
from pageObjects.companySignUpPage import companySignUpPage


class addEmployees(unittest.TestCase):
    baseURL = ReadConfig.getApplicationURL()
    DeptName = "Emp creation QA"
    DeptDescription = "Emp creation Software Testing"

    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    # username = worksheet["I2"].value
    username = worksheet["A2"].value
    password = ReadConfig.getPassword()

    workbook.close()

    logger = LogGen.loggen()

    def setUp(self):
        self.logger = LogGen.loggen()
        self.driver = webdriver.Firefox()  # Change to the appropriate driver
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)
        self.logger.info("****Opening URL****")
        self.driver.get(self.baseURL)

    def tearDown(self):
        self.driver.quit()

    @pytest.mark.run(order=1)
    @pytest.mark.regression
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_createEmployee_superAdmin(self):
        self.logger.info("****Started Create New Employee in Super Admin and Admin Account ****")
        first_name = randomGen.random_first_name()
        email = randomGen.random_email()
        phone_number = randomGen.random_phone_number()
        Emp_Id = randomGen.random_Emp_Id()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A8'] = first_name
        ws['B8'] = "personal" + email
        ws['D8'] = phone_number
        ws['C8'] = Emp_Id
        ws['E8'] = "emp" + email

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.aep.clickActive()
        time.sleep(2)
        self.aep.clickNewButton()
        self.aep.setFullname(first_name)

        self.aep.setEmail("emp" + email)
        self.aep.setPersonalEmail("personal" + email)

        self.aep.setPhoneNumber(phone_number)

        self.aep.setEmpId(Emp_Id)
        self.logger.info("**** Started Create  New Department in Super Admin Account****")
        self.aep.clickAddDeptButton()
        self.cp = ConfigurationPage(self.driver)
        self.cp.setDepartmentName("Department " + first_name)
        self.cp.setEnterDescription(self.DeptDescription)
        self.aep.clickDoneAddDept()
        time.sleep(3)
        act_Text = self.cp.Text_DeptCreatedSuccessful()
        if act_Text == "Department created successfully":
            assert True
            self.logger.info("********* Create Department Test is Passed ***********")

        else:
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_createDept.png")
            self.logger.error("********* Create Department Test is Failed ***********")
            self.driver.close()
            assert False

        self.logger.info("**** Started Create  New Division in Super Admin Account****")
        self.aep.clickAddDivisionButton()
        self.aep.setDivisionName("Division " + first_name)
        self.cp.setEnterDescription(self.DeptDescription)
        self.aep.clickDoneAddDept()
        time.sleep(3)
        act_Text = self.cp.Text_DivisionCreatedSuccessful()
        if act_Text == "Division created successfully":
            assert True
            self.logger.info("********* Create Division Test is Passed ***********")

        else:
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_createDept.png")
            self.logger.error("********* Create Division Test is Failed ***********")
            self.driver.close()
            assert False

        self.logger.info("**** Started Create  New Designation in Super Admin Account****")
        self.aep.clickAddDesignation()
        self.aep.setDivisionName("Designation " + first_name)
        self.cp.setEnterDescription(self.DeptDescription)
        self.aep.clickDoneAddDept()
        time.sleep(3)
        act_Text = self.cp.Text_DesignationCreatedSuccessful()
        if act_Text == "Designation created successfully":
            assert True
            self.logger.info("********* Create Designation Test is Passed ***********")
        else:
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_createDept.png")
            self.logger.error("********* Create Designation Test is Failed ***********")
            self.driver.close()
            assert False

        # time.sleep(2)
        self.aep.clickCountryDD()
        # time.sleep(3)
        self.sp = companySignUpPage(self.driver)
        self.sp.clickindia()
        self.sp.clickstatedd()
        self.sp.clickTelangana()
        self.sp.clickcitydd()
        self.sp.clickHyderabad()
        time.sleep(2)
        self.aep.clickAddButton()
        xpath = "//div[contains(text(), 'Employee created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            self.driver.quit()
            assert False


    @pytest.mark.run(order=2)
    @pytest.mark.regression
    def test_Employee_StatusAndRole(self):
        self.logger.info("****Started Create New Employee in Super Admin and Admin Account ****")
        first_name = randomGen.random_first_name()
        email = randomGen.random_email()
        phone_number = randomGen.random_phone_number()
        Emp_Id = randomGen.random_Emp_Id()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A11'] = first_name
        ws['B11'] = "personal" + email
        ws['D11'] = phone_number
        ws['C11'] = Emp_Id
        ws['E11'] = "emp" + email

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.aep.clickActive()
        time.sleep(2)
        self.aep.clickNewButton()
        self.aep.setFullname(first_name)

        self.aep.setEmail("emp" + email)
        self.aep.setPersonalEmail("personal" + email)

        self.aep.setPhoneNumber(phone_number)

        self.aep.setEmpId(Emp_Id)
        time.sleep(2)
        self.aep.ClickDD_Dept()
        self.aep.perform_keyboard_actions()
        self.aep.ClickDD_Division()
        self.aep.perform_keyboard_actions()
        self.aep.ClickDD_Designation()
        self.aep.perform_keyboard_actions()
        # time.sleep(2)
        self.aep.clickCountryDD()
        # time.sleep(3)
        self.sp = companySignUpPage(self.driver)
        self.sp.clickindia()
        self.sp.clickstatedd()
        self.sp.clickTelangana()
        self.sp.clickcitydd()
        self.sp.clickHyderabad()
        time.sleep(1)
        self.aep.clickAddButton()
        # time.sleep(3)

        xpath = "//div[contains(text(), 'Employee created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            self.driver.quit()
            assert False

        # if "Employee created successfully" in self.driver.page_source:
        #     self.logger.info("********** Acronym creation test is passed *********")
        #
        # else:
        #     # Log and take a screenshot
        #     self.logger.error("************** Acronym creation test is failed **********")
        #     self.driver.save_screenshot(".\\Screenshots\\" + "test_EmployeeCreated.png")
        #     assert False

        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        first_name = ws['A11'].value
        self.aep.clickActive()
        # time.sleep(2)
        self.aep.setActiveSearchField(first_name)

        element = self.driver.find_element(By.XPATH, "//span[contains(text(),'" + first_name + "')]")
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_ApproveSignedUpEmployee.png")
            self.driver.close()
            assert False

        element.click()
        self.aep.ClickEmployeeStatus()
        self.aep.ClickAdminStatus()
        self.aep.ClickGrantAdmin()
        xpath = "//div[contains(text(), '" + first_name + " is an admin now')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            self.driver.quit()
            assert False
        self.aep.ClickAdminStatus()
        self.aep.ClickEmployeesStatus()
        self.aep.ClickRemoveStatus()
        xpath = "//div[contains(text(), '" + first_name + " is removed as admin')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            self.driver.quit()
            assert False

        self.aep.ClickbuttonActive()
        self.aep.ClickbuttonDeactivate()
        self.aep.setreasonText("deleting to test the functionality")
        self.aep.ClickconfDeactivate()

        xpath = "//div[contains(text(),'Employee deactivated successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            self.driver.quit()
            assert False
        time.sleep(1)
        self.aep.setActiveSearchField(first_name)

        if "No search results found" in self.driver.page_source:
            self.logger.info("********** Acronym creation test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Acronym creation test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeed1.png")
            assert False


    if __name__ == '__main__':
        unittest.main(verbosity=2)
