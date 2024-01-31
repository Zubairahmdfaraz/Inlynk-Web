import time
import unittest

# from telnetlib import EC

import pytest
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from selenium.common import NoSuchElementException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

from pageObjects.LoginPage import LoginPage
from testCases.conftest import setup
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig
from pageObjects.companySignUpPage import companySignUpPage
from pageObjects.randomGen import randomGen
from selenium import webdriver
from openpyxl import workbook


class TestSignUp(unittest.TestCase):
    baseURL = ReadConfig.getApplicationURL()
    setSearchIndustryType = "Information Technology"
    password = ReadConfig.getPassword()

    def setUp(self):
        self.logger = LogGen.loggen()
        self.driver = webdriver.Chrome()  # Change to the appropriate driver
        self.driver.maximize_window()
        self.driver.implicitly_wait(20)
    def tearDown(self):
        self.driver.quit()

    @pytest.mark.run(order=1)
    @pytest.mark.smoke
    def test_SignUpwithValid(self):
        self.driver.get(self.baseURL)
        self.logger.info("******** Starting test_Sign Up with Valid ***********")
        self.logger.info("******** User is on Login page ***********")

        email = randomGen.random_email()
        first_name = randomGen.random_first_name()
        company_name = randomGen.random_company_name()
        phone_number = randomGen.random_phone_number()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A2'] = email
        ws['B2'] = first_name
        ws['C2'] = company_name
        ws['D2'] = phone_number

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.clickCompanysignupButton()
        self.logger.info("******** user is in company signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        self.sp.setCompanyName(company_name)

        self.sp.setSearchIndustryType(self.setSearchIndustryType)
        self.sp.selectCompany()
        self.sp.setContactName(first_name)
        self.sp.setEmail(email)
        self.sp.clickcountrydd()
        self.sp.clickindia()
        self.sp.clickstatedd()
        self.sp.clickTelangana()
        self.sp.clickcitydd()
        self.sp.clickHyderabad()
        self.sp.setPhone(phone_number)
        self.sp.setPassword(self.password)
        self.sp.setConfirmPassword(self.password)
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** user navigated to enter OTP page ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

# Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("https://yopmail.com/")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@id='login']")
        yopmail.send_keys(email + Keys.ENTER)
        time.sleep(1)
        iframeElement = self.driver.find_element(By.ID, "ifmail")
        self.driver.switch_to.frame(iframeElement)

        # here appears the captcha
        captcha_retries = 3
        attempts = 0
        captcha_found = False
        while attempts < captcha_retries:
            try:

                # If CAPTCHA is successfully filled, set captcha_found to True
                captcha_found = False  # Update this line accordingly

                # Verify the presence of specific text
                text_element = self.driver.find_element(By.XPATH, "//div[@class='b alc r_message']")
                if text_element.text == "Complete the CAPTCHA to continue":
                    captcha_found = True
                    break  # Exit the loop if expected text is found

            except NoSuchElementException:
                # No CAPTCHA found, continue with the rest of the steps
                captcha_found = False  # Update this line accordingly
                break

            except Exception as e:
                print(f"Error occurred: {e}")
                attempts += 1
                time.sleep(2)

            if captcha_found:
                print("Expected text found. Rerunning the test.")
                self.test_SignUpwithValid(setup)  # Rerun the test method

            else:
                print("Expected text not found. Continuing with the code.")
                # Continue with the remaining steps of your test script
                time.sleep(2)  # Add any necessary wait times or other steps

        # iframeElement = self.driver.find_element(By.ID, "ifmail")
        # self.driver.switch_to.frame(iframeElement)
        self.logger.info("******** Email Copied ***********")
        time.sleep(1)
        otp = self.driver.find_element(By.XPATH,
                                       "/html[1]/body[1]/main[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/h1[1]")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)
        getOTP = otp.text
        print(getOTP)
        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])

        self.sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info("******** Verifying the OTP ***********")
        self.sp.clickVerifyButton()
        self.sp.clickContinueToLogin()
        self.logger.info("******** Company Sign Up successful ***********")
        self.logger.info("******** Entering the sig up credentials for Login ***********")
        # Read data from specific cells
        email = ws['A2'].value

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickcreatePost()
        self.logger.info("******** Login successful ***********")
        act_Text = self.lp.newsFeedText()

        if act_Text == "Create News Feed":
            assert True
            self.logger.info("********* SignUp Test is Passed ***********")

        else:
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_SignUpwithValid.png")
            self.logger.error("********* SignUp Test is Failed ***********")
            self.driver.close()
            assert False

        time.sleep(3)
        self.driver.find_element(By.XPATH, "//div[@class='flexAutoRow alignCntr pdngHXS']").click()
        self.driver.close()

    if __name__ == '__main__':
        unittest.main(verbosity=2)